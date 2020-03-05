#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""=================================================
@Project -> File   ：mylearn -> Excel2PDF
@IDE    ：PyCharm
@Author ：Mr. toiler
@Date   ：03/02/2020 23:38 PM
@Desc   ：利用 openpyxl、reportlab 把excel文档转为PDF
=================================================="""
import openpyxl
import os

from reportlab.lib.units import inch, mm
from reportlab.lib.pagesizes import A3, A4, landscape


from utils import get_pdf_filename, PDFWriter


class Xlsx2PDF(object):
    COLUMN = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def __init__(self, filename, is_overwrite=False):
        self.xls_path, self.xls_name = os.path.split(os.path.abspath(filename))
        self.is_overwrite = is_overwrite
        self.rb: openpyxl.Workbook = openpyxl.load_workbook(filename, read_only=False)

    def convert(self, pdf_path=None):
        for rs in self.rb.worksheets[0:1]:
            filename = get_pdf_filename(xls_path=self.xls_path, xls_name=self.xls_name,
                                        sheet_name=rs.title, pdf_path=pdf_path, is_overwrite=self.is_overwrite)
            pdf_size, col_widths, row_heights, zoom, start_pos = self.get_size(rs=rs, pagesize=(420*mm, 297*mm))
            pdf_writer = PDFWriter(filename, pdf_size)

            print(pdf_size, (sum(col_widths), sum(row_heights)), zoom, start_pos)
            print(col_widths)
            print(row_heights)

            line_list, bg_rect_list, cell_rect_map = self.get_lines(rs, col_widths, row_heights, start_pos)

            print(line_list)
            print(bg_rect_list)
            print(cell_rect_map)

    def get_lines(self, rs,  widths,  heights, start_pos=(0, 0)):
        """得到绘制的表格线、背景框底色和坐标、实际表格的坐标"""
        n_cols = len(widths)
        n_rows = len(heights)
        line_list = list()
        bg_rect = list()
        merged_cell_rect = dict()
        pos = [start_pos[0], start_pos[1]]
        leave_lines, merged_cell_size = self.get_merged_cell_info(rs, widths, heights)
        print(leave_lines, merged_cell_size)
        # for n_row in range(n_rows-1, -1, -1):
        #     for n_col in range(0, n_cols):
        #         key = (n_row, n_col)
        #         xf: XF = rs.book.xf_list[rs.cell(n_row, n_col).xf_index]
        #         cell_size = (widths[n_col], heights[n_row])  # w, h
        #         leave_line_flag = leave_lines.get(key, [True, True, True, True])
        #         self.add_cell_lines(line_list, xf.border, leave_line_flag, pos, cell_size)
        #         bg: XFBackground = xf.background
        #         if bg.fill_pattern == 1:
        #             rgb = rs.book.colour_map.get(bg.pattern_colour_index)
        #             if rgb:
        #                 bg_rect.append((Color(rgb[0]/255, rgb[1]/255, rgb[2]/255), (pos[0], pos[1]), cell_size))
        #         # 把实际单元格的大小也一并返回去，合并和非合并的单元格都返回去
        #         if key in merged_cell_size.keys():
        #             s = merged_cell_size[key]
        #             merged_cell_rect[key] = ((pos[0], pos[1] + cell_size[1] - s[1]), s)
        #         else:
        #             merged_cell_rect[key] = ((pos[0], pos[1]), cell_size)
        #         pos[0] = pos[0] + cell_size[0]
        #     pos[0] = start_pos[0]
        #     pos[1] = pos[1] + cell_size[1]
        return line_list, bg_rect, merged_cell_rect

    @staticmethod
    def get_merged_cell_info(rs, widths, heights):
        """得到合并表格真正需要留下的边框， 合并的内部表格，不需要保留边框线"""
        leave_lines = dict()  # 合并单元格，内部的线就不需要了，需要删除
        merged_cell_size = dict()  # 合并单元格，真正的宽度和高度
        for cells in rs.merged_cells.ranges:
            row1, row2, col1, col2 = cells.min_row-1, cells.max_row, cells.min_col - 1, cells.max_col
            merged_cell_size[(row1, col1)] = (sum(widths[col1:col2]), sum(heights[row1:row2]))
            for n_row in range(row1, row2):
                for n_col in range(col1, col2):
                    border = (n_col == col1, n_row == row1, n_col == (col2 - 1), n_row == (row2 - 1))
                    leave_lines[(n_row, n_col)] = border
        return leave_lines, merged_cell_size

    @staticmethod
    def get_size(rs, pagesize=None, margin=0.04):
        if rs.max_column > len(Xlsx2PDF.COLUMN):
            raise Exception('column over {}.'.format(len(Xlsx2PDF.COLUMN)))

        dh_col = rs.column_dimensions
        dh_row = rs.row_dimensions
        # Excel中： 单位的换算：72磅 = 1 英寸，1 英寸 = 2.54  厘米。
        #  行高（单位磅）:          1毫米＝2.7682个单位；  1厘米＝27.682个单位； 1个单位＝0.3612 毫米。
        #  列宽（单位1/10英寸）:    1毫米＝0.4374个单位；  1厘米＝4.374个单位；  1个单位＝2.2862 毫米。
        # /2 这个是没有道理的，不知道，应怎么计算，不除比例不合适
        col_widths = [dh_col[Xlsx2PDF.COLUMN[n_col]].width * 2.2862/2 * mm for n_col in range(0, rs.max_column)]
        row_heights = [dh_row[n_row].height * 0.3612 * mm for n_row in range(1, rs.max_row + 1)]

        p_size = (pagesize[0] * (1 - margin), pagesize[1] * (1 - margin))  # 准备页边距留 4%， 1000 留 40点
        zoom1 = max(sum(col_widths), sum(row_heights)) / max(p_size)
        zoom2 = min(sum(col_widths), sum(row_heights)) / min(p_size)

        zoom = max(zoom1, zoom2)
        # if zoom < 1:
        #     zoom = 1

        col_widths = [c_w / zoom for c_w in col_widths]
        row_heights = [r_h / zoom for r_h in row_heights]

        # todo: 当一页无法放下时候，没有考虑，无论如何都会压到一页上，哪怕字体很小，以后添加字体很小的判断，保证可以看清楚
        s_w = sum(col_widths)
        s_h = sum(row_heights)
        if s_w > s_h:
            w, h = (max(pagesize), min(pagesize))
        else:
            w, h = (min(pagesize), max(pagesize))
        x = (w - s_w) / 2
        y = (h - s_h) / 2

        return (w, h), col_widths, row_heights, zoom, (x, y)


if __name__ == '__main__':
    import time
    t = time.time()
    print('start convert!')
    excel_file_path = r'c:/Users/zbd/PycharmProjects/mylearn/res/excel.xlsx'
    # excel_file_path = r'c:/Users/zbd/PycharmProjects/mylearn/res/整理.xls'
    # excel_file_path = r'd:/test.xlsx'

    xls2pdf = Xlsx2PDF(excel_file_path, True)
    xls2pdf.convert()
    print('time cost: {}'.format(time.time() - t))